#!/usr/bin/env python3
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
import os
import json

CONFIG_FILE = "schedule_rules.json"

def load_or_create_config():
    """加载配置文件，如果不存在则创建默认配置"""
    default_config = {
        "version": "2.0",
        "daily_keywords": ['血', '急', '凝', '服', '接', '细', '肿', '小', '白', '骨', '生', '免', 'P日', '抽', '早', '精', '科', '结', '公'],
        "exact_match_rules": {
            '前1': '中/前',
            '前2': '中/前',
            '急帮': '日/帮',
            '急帮休': '日/帮',
            '后1': '上午班/休',
            '后2': '上午班/休',
            '备': '休',
            '休': '休',
            '产假': '产假',
        },
        "triangle_rules": {
            'trigger_cells': ['前1', '前2', '后1', '后2', '后1/中', '后2/中'],
            'front_value': '值休',
            'back_value': '后夜'
        }
    }
    
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
            print(f"✅ 已加载配置文件: {CONFIG_FILE}")
            return config
        except Exception as e:
            print(f"⚠️ 配置文件读取失败，使用默认配置: {e}")
            return default_config
    else:
        # 创建默认配置文件
        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, ensure_ascii=False, indent=2)
            print(f"✅ 已创建默认配置文件: {CONFIG_FILE}")
        except Exception as e:
            print(f"⚠️ 创建配置文件失败: {e}")
        return default_config

def save_config(config):
    """保存配置到文件"""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"✅ 配置已保存到: {CONFIG_FILE}")
        return True
    except Exception as e:
        print(f"❌ 保存配置失败: {e}")
        return False

def show_menu():
    """显示主菜单"""
    print("\n" + "="*60)
    print("           排班自动替换工具 v2.0 - 规则配置菜单")
    print("="*60)
    print("  1. 直接处理Excel文件（使用当前规则）")
    print("  2. 查看当前所有规则")
    print("  3. 添加日班关键词")
    print("  4. 删除日班关键词")
    print("  5. 添加精确匹配规则")
    print("  6. 删除精确匹配规则")
    print("  7. 恢复默认规则")
    print("  8. 退出")
    print("="*60)
    choice = input("请选择操作 (1-8): ").strip()
    return choice

def view_rules(config):
    """查看当前规则"""
    print("\n" + "-"*60)
    print("【当前日班关键词】")
    keywords = config.get('daily_keywords', [])
    print(f"共 {len(keywords)} 个: {', '.join(keywords)}")
    
    print("\n【精确匹配规则】")
    rules = config.get('exact_match_rules', {})
    print(f"共 {len(rules)} 条:")
    for k, v in rules.items():
        print(f"  '{k}' → '{v}'")
    
    print("\n【三角形替换规则】")
    tri = config.get('triangle_rules', {})
    triggers = tri.get('trigger_cells', [])
    front = tri.get('front_value', '值休')
    back = tri.get('back_value', '后夜')
    print(f"  触发单元格: {', '.join(triggers)}")
    print(f"  前1/前2 后面的△ → {front}")
    print(f"  其他触发单元格后面的△ → {back}")
    print("-"*60)
    input("\n按回车键继续...")

def add_daily_keyword(config):
    """添加日班关键词"""
    keyword = input("\n请输入要添加的日班关键词: ").strip()
    if not keyword:
        print("❌ 关键词不能为空")
        return
    
    if keyword not in config['daily_keywords']:
        config['daily_keywords'].append(keyword)
        if save_config(config):
            print(f"✅ 已添加关键词: '{keyword}'")
    else:
        print(f"⚠️ 关键词 '{keyword}' 已存在")
    input("\n按回车键继续...")

def delete_daily_keyword(config):
    """删除日班关键词"""
    print("\n当前日班关键词列表:")
    for i, kw in enumerate(config['daily_keywords'], 1):
        print(f"  {i}. {kw}")
    
    try:
        idx = int(input("\n请输入要删除的序号: ").strip()) - 1
        if 0 <= idx < len(config['daily_keywords']):
            kw = config['daily_keywords'].pop(idx)
            if save_config(config):
                print(f"✅ 已删除关键词: '{kw}'")
        else:
            print("❌ 序号无效")
    except ValueError:
        print("❌ 请输入有效的数字")
    input("\n按回车键继续...")

def add_exact_rule(config):
    """添加精确匹配规则"""
    key = input("\n请输入要匹配的内容: ").strip()
    if not key:
        print("❌ 匹配内容不能为空")
        input("\n按回车键继续...")
        return
    
    value = input("请输入替换后的内容: ").strip()
    if not value:
        print("❌ 替换内容不能为空")
        input("\n按回车键继续...")
        return
    
    config['exact_match_rules'][key] = value
    if save_config(config):
        print(f"✅ 已添加规则: '{key}' → '{value}'")
    input("\n按回车键继续...")

def delete_exact_rule(config):
    """删除精确匹配规则"""
    print("\n当前精确匹配规则列表:")
    rules = list(config['exact_match_rules'].items())
    for i, (k, v) in enumerate(rules, 1):
        print(f"  {i}. '{k}' → '{v}'")
    
    try:
        idx = int(input("\n请输入要删除的序号: ").strip()) - 1
        if 0 <= idx < len(rules):
            key, value = rules[idx]
            del config['exact_match_rules'][key]
            if save_config(config):
                print(f"✅ 已删除规则: '{key}' → '{value}'")
        else:
            print("❌ 序号无效")
    except ValueError:
        print("❌ 请输入有效的数字")
    input("\n按回车键继续...")

def reset_default(config):
    """恢复默认规则"""
    confirm = input("\n⚠️ 确定要恢复所有默认规则吗？这将清空您的自定义配置！(y/N): ").strip().lower()
    if confirm == 'y' or confirm == 'yes':
        config['daily_keywords'] = ['血', '急', '凝', '服', '接', '细', '肿', '小', '白', '骨', '生', '免', 'P日', '抽', '早', '精', '科', '结', '公']
        config['exact_match_rules'] = {
            '前1': '中/前', '前2': '中/前', '急帮': '日/帮', '急帮休': '日/帮',
            '后1': '上午班/休', '后2': '上午班/休', '备': '休', '休': '休', '产假': '产假',
        }
        if save_config(config):
            print("✅ 已恢复默认规则")
    else:
        print("已取消")
    input("\n按回车键继续...")

def modify_schedule(input_path, output_path, config):
    """处理Excel文件"""
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    daily_words = set(config.get('daily_keywords', []))
    exact_match_replace = config.get('exact_match_rules', {}).copy()
    
    # 日班关键词也加入精确匹配（单个关键词直接替换为日班）
    for word in daily_words:
        if word not in exact_match_replace:
            exact_match_replace[word] = '日班'
    
    triangle_rules = config.get('triangle_rules', {})
    trigger_cells = triangle_rules.get('trigger_cells', ['前1', '前2', '后1', '后2', '后1/中', '后2/中'])
    front_value = triangle_rules.get('front_value', '值休')
    back_value = triangle_rules.get('back_value', '后夜')

    wb = load_workbook(input_path)
    modified_count = 0
    highlight_count = 0
    triangle_changed = 0
    replaced_cells = set()

    # 第一步：处理三角形替换
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_text = str(cell.value).strip()
                if cell_text in trigger_cells:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if next_cell.value is None:
                        continue
                    next_text = str(next_cell.value).strip()
                    if next_text == '△' or next_text == '▲':
                        if cell_text in ['前1', '前2']:
                            next_cell.value = front_value
                        else:
                            next_cell.value = back_value
                        replaced_cells.add((sheet_name, next_cell.row, next_cell.column))
                        triangle_changed += 1
                        modified_count += 1

    # 第二步：处理其他替换规则
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_text = str(cell.value).strip()
                if (sheet_name, cell.row, cell.column) in replaced_cells:
                    continue
                if '/中' in cell_text or cell_text.startswith('中') or cell_text.endswith('中'):
                    cell.value = '日班'
                    modified_count += 1
                    continue
                if len(cell_text) == 2 and cell_text.endswith('休'):
                    cell.value = '上午班/休'
                    modified_count += 1
                    continue
                if len(cell_text) == 2 and cell_text.startswith('休'):
                    cell.value = '休/下午班'
                    modified_count += 1
                    continue
                if cell_text in exact_match_replace:
                    cell.value = exact_match_replace[cell_text]
                    modified_count += 1
                    continue
                count = 0
                for word in daily_words:
                    if word in cell_text:
                        count += 1
                if count >= 2:
                    cell.value = '日班'
                    modified_count += 1
                    continue
                if cell.row > 1 and cell_text.strip() != '':
                    cell.fill = yellow_fill
                    highlight_count += 1

    wb.save(output_path)
    print(f"\n✅ 修改完成!")
    print(f"  替换单元格数: {modified_count}")
    print(f"  其中三角形替换: {triangle_changed}")
    print(f"  标黄单元格数: {highlight_count}")
    print(f"  输出文件: {output_path}")
    return modified_count, highlight_count

def main():
    print("="*60)
    print("           排班自动替换工具 v2.0")
    print("="*60)
    
    # 加载配置
    config = load_or_create_config()
    
    # 如果命令行直接传入了文件参数，直接处理，不显示菜单
    if len(sys.argv) >= 2:
        input_path = sys.argv[1]
        if len(sys.argv) == 3:
            output_path = sys.argv[2]
        else:
            name, ext = os.path.splitext(input_path)
            output_path = f"{name}_modified{ext}"
        modify_schedule(input_path, output_path, config)
        return
    
    # 交互式菜单
    while True:
        choice = show_menu()
        
        if choice == '1':
            input_path = input("\n请输入Excel文件路径: ").strip()
            if not input_path:
                print("❌ 路径不能为空")
                input("\n按回车键继续...")
                continue
            if not os.path.exists(input_path):
                print(f"❌ 文件不存在: {input_path}")
                input("\n按回车键继续...")
                continue
            
            name, ext = os.path.splitext(input_path)
            output_path = f"{name}_modified{ext}"
            modify_schedule(input_path, output_path, config)
            input("\n按回车键继续...")
            
        elif choice == '2':
            view_rules(config)
            
        elif choice == '3':
            add_daily_keyword(config)
            
        elif choice == '4':
            delete_daily_keyword(config)
            
        elif choice == '5':
            add_exact_rule(config)
            
        elif choice == '6':
            delete_exact_rule(config)
            
        elif choice == '7':
            reset_default(config)
            
        elif choice == '8':
            print("\n👋 再见！")
            break
            
        else:
            print("\n❌ 无效的选择，请重新输入")

if __name__ == "__main__":
    main()
