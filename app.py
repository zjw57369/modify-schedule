#!/usr/bin/env python3
from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import json
import io
import tempfile

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

CONFIG_FILE = "schedule_rules.json"
UPLOAD_FOLDER = tempfile.mkdtemp()

# 默认配置
DEFAULT_CONFIG = {
    "version": "2.0-web",
    "daily_keywords": ['血', '急', '凝', '服', '接', '细', '肿', '小', '白', '骨', '生', '免', 'P日', '抽', '早', '精', '科', '结', '公'],
    "exact_match_rules": {
        '前1': '中/前', '前2': '中/前', '急帮': '日/帮', '急帮休': '日/帮',
        '后1': '上午班/休', '后2': '上午班/休', '备': '休', '休': '休', '产假': '产假',
    },
    "triangle_rules": {
        'trigger_cells': ['前1', '前2', '后1', '后2', '后1/中', '后2/中'],
        'front_value': '值休',
        'back_value': '后夜'
    }
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return DEFAULT_CONFIG.copy()

def save_config(config):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False

def process_excel(file_content, config):
    """处理Excel文件，返回处理后的文件内容"""
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    daily_words = set(config.get('daily_keywords', []))
    exact_match_replace = config.get('exact_match_rules', {}).copy()
    
    for word in daily_words:
        if word not in exact_match_replace:
            exact_match_replace[word] = '日班'
    
    triangle_rules = config.get('triangle_rules', {})
    trigger_cells = triangle_rules.get('trigger_cells', ['前1', '前2', '后1', '后2', '后1/中', '后2/中'])
    front_value = triangle_rules.get('front_value', '值休')
    back_value = triangle_rules.get('back_value', '后夜')

    wb = load_workbook(io.BytesIO(file_content))
    modified_count = 0
    highlight_count = 0
    triangle_changed = 0
    replaced_cells = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_text = str(cell.value).strip()
                if cell_text in trigger_cells:
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if next_cell.value is None:
                        continue
                    next_text = str(next_cell.value).strip()
                    if next_text == '△' or next_text == '▲':
                        if cell_text in ['前1', '前2']:
                            next_cell.value = front_value
                        else:
                            next_cell.value = back_value
                        replaced_cells.add((sheet_name, next_cell.row, next_cell.column))
                        triangle_changed += 1
                        modified_count += 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_text = str(cell.value).strip()
                if (sheet_name, cell.row, cell.column) in replaced_cells:
                    continue
                if '/中' in cell_text or cell_text.startswith('中') or cell_text.endswith('中'):
                    cell.value = '日班'
                    modified_count += 1
                    continue
                if len(cell_text) == 2 and cell_text.endswith('休'):
                    cell.value = '上午班/休'
                    modified_count += 1
                    continue
                if len(cell_text) == 2 and cell_text.startswith('休'):
                    cell.value = '休/下午班'
                    modified_count += 1
                    continue
                if cell_text in exact_match_replace:
                    cell.value = exact_match_replace[cell_text]
                    modified_count += 1
                    continue
                count = 0
                for word in daily_words:
                    if word in cell_text:
                        count += 1
                if count >= 2:
                    cell.value = '日班'
                    modified_count += 1
                    continue
                if cell.row > 1 and cell_text.strip() != '':
                    cell.fill = yellow_fill
                    highlight_count += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue(), {
        'modified': modified_count,
        'highlighted': highlight_count,
        'triangle': triangle_changed
    }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/config', methods=['GET'])
def get_config():
    return jsonify(load_config())

@app.route('/api/config', methods=['POST'])
def update_config():
    config = request.json
    if save_config(config):
        return jsonify({'success': True, 'message': '配置已保存'})
    return jsonify({'success': False, 'message': '保存失败'})

@app.route('/api/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有上传文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '只支持.xlsx或.xls格式的Excel文件'})
    
    try:
        file_content = file.read()
        config = load_config()
        result_content, stats = process_excel(file_content, config)
        
        output_filename = os.path.splitext(file.filename)[0] + '_modified.xlsx'
        
        return send_file(
            io.BytesIO(result_content),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_filename
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'处理失败: {str(e)}'})

if __name__ == '__main__':
    print("="*60)
    print("  排班自动替换工具 - 网页版 v2.0")
    print("="*60)
    print("  请在浏览器中打开: http://localhost:5000")
    print("  按 Ctrl+C 停止服务")
    print("="*60)
    print()
    app.run(host='0.0.0.0', port=5000, debug=False)
